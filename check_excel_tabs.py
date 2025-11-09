"""
Check which tabs in the Excel file are blank or sparse
"""
import openpyxl
from pathlib import Path

def check_excel_tabs(filepath):
    """Check all tabs in Excel file for content"""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    
    print(f"\n{'='*80}")
    print(f"EXCEL TAB ANALYSIS: {Path(filepath).name}")
    print(f"{'='*80}\n")
    
    results = {}
    
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        
        # Count cells with content
        total_cells = 0
        filled_cells = 0
        
        for row in sheet.iter_rows():
            for cell in row:
                total_cells += 1
                if cell.value is not None and str(cell.value).strip():
                    filled_cells += 1
        
        fill_rate = (filled_cells / total_cells * 100) if total_cells > 0 else 0
        
        status = "✅ GOOD" if filled_cells > 50 else "⚠️  SPARSE" if filled_cells > 0 else "❌ EMPTY"
        
        results[sheet_name] = {
            'filled_cells': filled_cells,
            'total_cells': total_cells,
            'fill_rate': fill_rate,
            'status': status
        }
        
        print(f"{status} {sheet_name:30s} - {filled_cells:>4} / {total_cells:>6} cells ({fill_rate:>5.1f}%)")
    
    print(f"\n{'='*80}")
    print(f"SUMMARY:")
    empty = [name for name, data in results.items() if data['filled_cells'] == 0]
    sparse = [name for name, data in results.items() if 0 < data['filled_cells'] <= 50]
    
    if empty:
        print(f"\n❌ EMPTY TABS ({len(empty)}):")
        for name in empty:
            print(f"   - {name}")
    
    if sparse:
        print(f"\n⚠️  SPARSE TABS ({len(sparse)}):")
        for name in sparse:
            print(f"   - {name}")
    
    print(f"\n{'='*80}\n")
    
    return results

if __name__ == "__main__":
    excel_file = "outputs/NVDA_Comprehensive_Model_20251107_1210.xlsx"
    check_excel_tabs(excel_file)
